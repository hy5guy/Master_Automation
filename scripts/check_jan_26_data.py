#!/usr/bin/env python3
import openpyxl
import sys

if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\carucci_r\OneDrive - City of Hackensack\Shared Folder\Compstat\Contributions\Detectives\detectives_monthly.xlsx")

print("=" * 80)
print("CHECKING 26-Jan COLUMN DATA")
print("=" * 80)

# Check MoM sheet
ws = wb['MoM']
headers = [cell.value for cell in ws[1]]

# Find 26-Jan column
jan_26_col = None
for idx, header in enumerate(headers):
    if header == "26-Jan":
        jan_26_col = idx + 1  # Excel is 1-indexed
        print(f"\nFound '26-Jan' at column index {jan_26_col} (Excel column {chr(64 + jan_26_col)})")
        break

if jan_26_col:
    print(f"\nSample data from '26-Jan' column (first 20 rows):")
    print("-" * 80)
    
    has_data = False
    for row_idx in range(2, 22):  # Check rows 2-21
        cell_value = ws.cell(row_idx, jan_26_col).value
        row_label = ws.cell(row_idx, 1).value
        print(f"  Row {row_idx} ({row_label}): {cell_value}")
        if cell_value is not None and cell_value != 0:
            has_data = True
    
    if not has_data:
        print("\n*** All values are NULL or 0 in 26-Jan column! ***")
    else:
        print("\n26-Jan column contains data.")
else:
    print("\n*** '26-Jan' column NOT FOUND in headers! ***")

print("\n" + "=" * 80)
