#!/usr/bin/env python3
import openpyxl
import sys

# Force UTF-8 output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\carucci_r\OneDrive - City of Hackensack\Shared Folder\Compstat\Contributions\Detectives\detectives_monthly.xlsx")

print("CHECKING CRITICAL SHEETS AND TABLES")
print("=" * 60)

# Check MoM sheet
if 'MoM' in wb.sheetnames:
    ws = wb['MoM']
    print("\n[OK] Sheet 'MoM' EXISTS")
    print(f"  Tables: {list(ws.tables.keys())}")
else:
    print("\n[MISS] Sheet 'MoM' NOT FOUND")

# Check CCD sheet
if 'CCD' in wb.sheetnames:
    ws = wb['CCD']
    print("\n[OK] Sheet 'CCD' EXISTS")
    print(f"  Tables: {list(ws.tables.keys())}")
else:
    print("\n[MISS] Sheet 'CCD' NOT FOUND")

# Check 26_JAN sheet
if '26_JAN' in wb.sheetnames:
    ws = wb['26_JAN']
    print("\n[OK] Sheet '26_JAN' EXISTS")
    print(f"  Tables: {list(ws.tables.keys())}")
else:
    print("\n[MISS] Sheet '26_JAN' NOT FOUND")

print("\n" + "=" * 60)
