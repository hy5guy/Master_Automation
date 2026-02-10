#!/usr/bin/env python3
"""
Fix WG2 column in summons_powerbi_latest.xlsx by copying from WG2_ASSIGN or Assignment Master.

This script:
1. Loads the staging workbook
2. For rows where WG2 is null:
   - First tries WG2_ASSIGN if it exists
   - Then tries Assignment Master lookup
3. Updates the workbook
"""

import pandas as pd
from pathlib import Path
import sys

def normalize_badge(badge):
    """Normalize badge to 4-digit string"""
    if pd.isna(badge):
        return ""
    badge_str = str(badge).strip()
    # Extract digits only
    digits = ''.join(c for c in badge_str if c.isdigit())
    # Pad to 4 digits
    if digits:
        return digits.zfill(4)
    return ""

def main():
    # Paths
    staging_file = Path(r"C:\Users\carucci_r\OneDrive - City of Hackensack\03_Staging\Summons\summons_powerbi_latest.xlsx")
    assignment_master = Path(r"C:\Users\carucci_r\OneDrive - City of Hackensack\09_Reference\Personnel\Assignment_Master_V2.csv")
    
    print("=" * 80)
    print("FIXING WG2 COLUMN IN SUMMONS DATA")
    print("=" * 80)
    
    # Load staging workbook
    if not staging_file.exists():
        print(f"❌ Staging file not found: {staging_file}")
        return 1
    
    print(f"\n✓ Loading staging workbook...")
    df = pd.read_excel(staging_file, sheet_name="Summons_Data")
    print(f"  Total rows: {len(df):,}")
    
    # Count rows needing fix
    wg2_null = df["WG2"].isna().sum()
    print(f"  Rows with null WG2: {wg2_null:,}")
    
    # Load Assignment Master
    print(f"\n✓ Loading Assignment Master...")
    am = pd.read_csv(assignment_master)
    
    # Normalize badge numbers in Assignment Master
    am["BADGE_NORMALIZED"] = am["PADDED_BADGE_NUMBER"].apply(normalize_badge)
    am_dict = dict(zip(am["BADGE_NORMALIZED"], am["WG2"]))
    
    # Create a copy for updates
    df_updated = df.copy()
    fixed_count = 0
    
    # Fix 1: Copy from WG2_ASSIGN if available
    if "WG2_ASSIGN" in df.columns:
        mask = df_updated["WG2"].isna() & df_updated["WG2_ASSIGN"].notna()
        count_from_assign = mask.sum()
        if count_from_assign > 0:
            df_updated.loc[mask, "WG2"] = df_updated.loc[mask, "WG2_ASSIGN"]
            fixed_count += count_from_assign
            print(f"\n✓ Fixed {count_from_assign:,} rows from WG2_ASSIGN")
    
    # Fix 2: Lookup from Assignment Master
    mask = df_updated["WG2"].isna()
    remaining = mask.sum()
    
    if remaining > 0:
        print(f"\n✓ Fixing {remaining:,} remaining rows from Assignment Master...")
        
        # Normalize badge numbers in summons data
        df_updated["BADGE_NORMALIZED"] = df_updated["PADDED_BADGE_NUMBER"].apply(normalize_badge)
        
        # Lookup WG2 from Assignment Master
        df_updated.loc[mask, "WG2"] = df_updated.loc[mask, "BADGE_NORMALIZED"].map(am_dict)
        
        fixed_from_am = df_updated.loc[mask, "WG2"].notna().sum()
        fixed_count += fixed_from_am
        print(f"  Fixed {fixed_from_am:,} rows from Assignment Master")
        
        # Remove temporary column
        df_updated = df_updated.drop(columns=["BADGE_NORMALIZED"])
    
    # Check final status
    wg2_still_null = df_updated["WG2"].isna().sum()
    wg2_has_value = len(df_updated) - wg2_still_null
    
    print(f"\n📊 Final Status:")
    print(f"  WG2 has value: {wg2_has_value:,}")
    print(f"  WG2 still null: {wg2_still_null:,}")
    print(f"  Total fixed: {fixed_count:,}")
    
    if wg2_still_null > 0:
        print(f"\n⚠️  {wg2_still_null:,} rows still have null WG2")
        print("   These may be badges not in Assignment Master or with null WG2 in AM")
    
    # Save updated workbook (fast method - write only WG2 column update)
    if fixed_count > 0:
        print(f"\n💾 Saving updated workbook...")
        print("   (This may take a minute for large files...)")
        
        # Read existing workbook
        from openpyxl import load_workbook
        wb = load_workbook(staging_file, read_only=False)
        ws = wb["Summons_Data"]
        
        # Find WG2 column
        wg2_col = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value == "WG2":
                wg2_col = col_idx
                break
        
        if wg2_col:
            # Update only WG2 column (much faster than rewriting entire sheet)
            print(f"   Updating WG2 column (column {wg2_col})...")
            for idx, value in enumerate(df_updated["WG2"], start=2):
                if idx % 50000 == 0:
                    print(f"   Progress: {idx:,} / {len(df_updated):,} rows")
                if pd.notna(value) and value != "":
                    ws.cell(row=idx, column=wg2_col, value=str(value))
                else:
                    ws.cell(row=idx, column=wg2_col, value=None)
        
        print("   Saving file...")
        wb.save(staging_file)
        wb.close()
        print(f"✓ Saved: {staging_file.name}")
    else:
        print(f"\n✓ No changes needed")
    
    print("\n" + "=" * 80)
    print("DONE!")
    print("=" * 80)
    print("\nNext steps:")
    print("1. Refresh your Power BI queries")
    print("2. The Top 5 Moving and Parking queries should now work correctly")
    print("=" * 80)
    
    return 0

if __name__ == "__main__":
    sys.exit(main())

