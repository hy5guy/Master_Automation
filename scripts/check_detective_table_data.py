#!/usr/bin/env python3
import openpyxl
import sys

# Force UTF-8 output
if sys.stdout.encoding != 'utf-8':
    sys.stdout.reconfigure(encoding='utf-8')

wb = openpyxl.load_workbook(r"C:\Users\carucci_r\OneDrive - City of Hackensack\Shared Folder\Compstat\Contributions\Detectives\detectives_monthly.xlsx")

print("=" * 80)
print("DETECTIVE WORKBOOK - DETAILED TABLE ANALYSIS")
print("=" * 80)

# Check MoM sheet and _mom_det table
print("\n1. MoM SHEET - _mom_det TABLE")
print("-" * 80)
if 'MoM' in wb.sheetnames:
    ws = wb['MoM']
    if '_mom_det' in ws.tables:
        table = ws.tables['_mom_det']
        print(f"  Table Name: {table.name}")
        print(f"  Table Range: {table.ref}")
        
        # Get the actual cell range
        min_row = ws[table.ref.split(':')[0]].row
        max_row = ws[table.ref.split(':')[1]].row if ':' in table.ref else min_row
        
        # Get headers (first row of table)
        headers = []
        for cell in ws[min_row]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"  Headers ({len(headers)} columns): {headers[:5]}...{headers[-3:]}")
        print(f"  Data rows: {max_row - min_row}")
        
        # Check first data row
        if max_row > min_row:
            first_data = []
            for cell in ws[min_row + 1]:
                first_data.append(cell.value)
            print(f"  First data row: {first_data[:3]}")
    else:
        print("  ERROR: _mom_det table not found!")
else:
    print("  ERROR: MoM sheet not found!")

# Check CCD sheet and _CCD_MOM table
print("\n2. CCD SHEET - _CCD_MOM TABLE")
print("-" * 80)
if 'CCD' in wb.sheetnames:
    ws = wb['CCD']
    if '_CCD_MOM' in ws.tables:
        table = ws.tables['_CCD_MOM']
        print(f"  Table Name: {table.name}")
        print(f"  Table Range: {table.ref}")
        
        # Get the actual cell range
        min_row = ws[table.ref.split(':')[0]].row
        max_row = ws[table.ref.split(':')[1]].row if ':' in table.ref else min_row
        
        # Get headers
        headers = []
        for cell in ws[min_row]:
            if cell.value:
                headers.append(cell.value)
        
        print(f"  Headers ({len(headers)} columns): {headers}")
        print(f"  Data rows: {max_row - min_row}")
        
        # Check all data rows for this table
        if max_row > min_row:
            print(f"\n  ALL ROWS:")
            for row_idx in range(min_row + 1, max_row + 1):
                row_values = [ws.cell(row_idx, col).value for col in range(1, len(headers) + 1)]
                print(f"    Row {row_idx - min_row}: {row_values[:3]}...")
    else:
        print("  ERROR: _CCD_MOM table not found!")
else:
    print("  ERROR: CCD sheet not found!")

print("\n" + "=" * 80)
