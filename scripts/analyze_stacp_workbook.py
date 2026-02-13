#!/usr/bin/env python3
"""
Analyze STACP.xlsm workbook structure
"""
import openpyxl
from pathlib import Path

# Load workbook
wb_path = r"C:\Users\carucci_r\OneDrive - City of Hackensack\Shared Folder\Compstat\Contributions\STACP\STACP.xlsm"
wb = openpyxl.load_workbook(wb_path, data_only=True)

print("=" * 80)
print("STACP WORKBOOK ANALYSIS")
print("=" * 80)

# Analyze MoMTotals sheet
print("\n1. MoMTotals Sheet Analysis")
print("-" * 80)
ws = wb['MoMTotals']

# Get headers
headers = [cell.value for cell in ws[1]]
print(f"Total columns: {len(headers)}")

# Find date columns
date_cols = []
for i, h in enumerate(headers):
    if h and isinstance(h, str) and '-' in h:
        date_cols.append((i+1, h))

print(f"\nDate columns found: {len(date_cols)}")
print("\nColumn Index | Header")
print("-" * 40)
for idx, header in date_cols:
    print(f"  Col {idx:3d}    | {header}")

# Check for data in first few rows
print("\n2. Sample Data Check (First 3 data rows)")
print("-" * 80)
for row_idx in range(2, 5):  # Rows 2-4
    row_data = [cell.value for cell in ws[row_idx]]
    tracked_item = row_data[0] if row_data else None
    print(f"\nRow {row_idx}: {tracked_item}")
    
    # Show values for date columns
    for col_idx, col_name in date_cols[:5]:  # First 5 date columns only
        val = row_data[col_idx-1] if col_idx <= len(row_data) else None
        print(f"  {col_name}: {val}")

# Check specific month sheets
print("\n3. Individual Month Sheet Check")
print("-" * 80)
target_sheets = ['25_JAN', '25_FEB', '25_MAR', '25_DEC', '26_JAN']

for sheet_name in target_sheets:
    if sheet_name in wb.sheetnames:
        ws_month = wb[sheet_name]
        # Check if sheet has data
        row_count = 0
        for row in ws_month.iter_rows(min_row=2, max_row=100):
            if any(cell.value for cell in row):
                row_count += 1
        print(f"  {sheet_name}: {row_count} rows with data")
    else:
        print(f"  {sheet_name}: NOT FOUND")

# Check MoMTotals formula or values
print("\n4. MoMTotals Data Source Check")
print("-" * 80)
print("Checking if MoMTotals contains formulas or static values...")

# Get cell A2 (first data cell)
cell_a2 = ws['A2']
print(f"Cell A2 value: {cell_a2.value}")
print(f"Cell A2 has formula: {cell_a2.data_type == 'f'}")

# Check a date column cell
if date_cols:
    first_date_col_letter = openpyxl.utils.get_column_letter(date_cols[0][0])
    test_cell = ws[f'{first_date_col_letter}2']
    print(f"\nCell {first_date_col_letter}2 ({date_cols[0][1]}) value: {test_cell.value}")
    print(f"Cell {first_date_col_letter}2 has formula: {test_cell.data_type == 'f'}")

print("\n" + "=" * 80)
print("Analysis complete!")
