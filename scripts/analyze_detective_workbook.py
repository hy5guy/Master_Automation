#!/usr/bin/env python3
"""
Analyze Detective workbook structure - sheets, tables, and data
"""
import openpyxl
from openpyxl.utils import get_column_letter

# Load workbook
wb_path = r"C:\Users\carucci_r\OneDrive - City of Hackensack\Shared Folder\Compstat\Contributions\Detectives\detectives_monthly.xlsx"
wb = openpyxl.load_workbook(wb_path, data_only=True)

print("=" * 80)
print("DETECTIVE WORKBOOK STRUCTURE ANALYSIS")
print("=" * 80)

# List all sheets
print("\n1. SHEETS IN WORKBOOK")
print("-" * 80)
for i, sheet_name in enumerate(wb.sheetnames):
    print(f"  {i+1}. {sheet_name}")

# Check for tables in each sheet
print("\n2. TABLES IN WORKBOOK")
print("-" * 80)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, '_tables') and ws._tables:
        table_names = [t.name for t in ws._tables]
        print(f"  Sheet '{sheet_name}':")
        for table_name in table_names:
            print(f"    - Table: {table_name}")
    else:
        # Check if it's a named sheet that might be treated as data
        print(f"  Sheet '{sheet_name}': No named tables (might be plain data)")

# Check MoMTotals sheet specifically (if it exists)
print("\n3. MoMTotals SHEET ANALYSIS")
print("-" * 80)
if 'MoMTotals' in wb.sheetnames:
    ws = wb['MoMTotals']
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"  Total columns: {len(headers)}")
    print(f"  Headers (first 15): {headers[:15]}")
    
    # Check for tables
    if hasattr(ws, '_tables') and ws._tables:
        table_names = [t.name for t in ws._tables]
        print(f"  Named tables: {table_names}")
    else:
        print("  No named tables (data might be in plain sheet)")
    
    # Check first data row
    row2 = [cell.value for cell in ws[2]]
    print(f"  First data row (A2): {row2[0] if row2 else 'Empty'}")
else:
    print("  MoMTotals sheet NOT FOUND")

# Check for _mom_det and _CCD_MOM specifically
print("\n4. LOOKING FOR SPECIFIC TABLES")
print("-" * 80)
target_tables = ['_mom_det', '_CCD_MOM', 'mom_det', 'CCD_MOM']
found_tables = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    if hasattr(ws, '_tables') and ws._tables:
        for table in ws._tables:
            if table.name in target_tables or table.name.lower() in [t.lower() for t in target_tables]:
                found_tables.append((sheet_name, table.name, table.ref))
                print(f"  ✓ Found: '{table.name}' in sheet '{sheet_name}' (Range: {table.ref})")

if not found_tables:
    print("  ✗ None of the target tables found!")
    print(f"  Looking for: {', '.join(target_tables)}")

# Check monthly sheets (26_JAN, 26_FEB, etc.)
print("\n5. MONTHLY SHEETS (2026)")
print("-" * 80)
monthly_sheets = [s for s in wb.sheetnames if s.startswith('26_')]
if monthly_sheets:
    print(f"  Found {len(monthly_sheets)} monthly sheets:")
    for sheet_name in sorted(monthly_sheets):
        ws = wb[sheet_name]
        table_count = len(ws._tables) if hasattr(ws, '_tables') and ws._tables else 0
        table_names = [t.name for t in ws._tables] if hasattr(ws, '_tables') and ws._tables else []
        print(f"    - {sheet_name}: {table_count} tables {table_names}")
else:
    print("  No monthly sheets found (26_JAN, 26_FEB, etc.)")

# Sample data check
print("\n6. SAMPLE DATA CHECK")
print("-" * 80)
if found_tables:
    sheet_name, table_name, table_ref = found_tables[0]
    ws = wb[sheet_name]
    print(f"  Checking table: {table_name} in sheet: {sheet_name}")
    
    # Get first 3 rows
    headers = [cell.value for cell in ws[1]]
    print(f"  Headers: {headers[:10]}")
    
    for row_idx in range(2, 5):
        row_data = [cell.value for cell in ws[row_idx]]
        print(f"  Row {row_idx}: {row_data[:5] if len(row_data) >= 5 else row_data}")

print("\n" + "=" * 80)
print("Analysis complete!")
