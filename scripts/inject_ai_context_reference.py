"""
inject_ai_context_reference.py — Inject AI_Context_Reference worksheet into HPD shared workbooks.

v2: Zip-level injection — never loads/saves the target workbook through openpyxl.
    Analysis uses read_only=True. Sheet is built in a temp workbook, then its XML
    and styles are surgically merged into the target zip. This preserves all existing
    data validations, conditional formatting, VBA, and XML extensions.

Usage:
    python scripts/inject_ai_context_reference.py                # All Tier 1 workbooks
    python scripts/inject_ai_context_reference.py --tier2        # Include Tier 2
    python scripts/inject_ai_context_reference.py --dry-run      # Preview only
    python scripts/inject_ai_context_reference.py --workbook CSB  # Single workbook
    python scripts/inject_ai_context_reference.py --verbose       # Detailed logging
    python scripts/inject_ai_context_reference.py --no-backup     # Skip backup (not recommended)
"""

import argparse
import io
import json
import os
import re
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------
ONEDRIVE_ROOT = Path(os.path.expanduser("~")) / "OneDrive - City of Hackensack"
CONTRIBUTIONS_DIR = ONEDRIVE_ROOT / "Shared Folder" / "Compstat" / "Contributions"
REPO_ROOT = ONEDRIVE_ROOT / "06_Workspace_Management"

# ---------------------------------------------------------------------------
# HPD color palette
# ---------------------------------------------------------------------------
NAVY = "1A2744"
GOLD = "C8A84B"
LIGHT_GOLD = "F5F0E0"
LIGHT_BLUE = "D6E4F0"
GRAY_ALT = "F2F2F2"
GRAY_TEXT = "666666"
LINK_BLUE = "0563C1"
WHITE = "FFFFFF"
BLACK = "000000"

# ---------------------------------------------------------------------------
# Workbook inventory
# ---------------------------------------------------------------------------
TIER1_WORKBOOKS = [
    {"id": 1, "name": "csb_monthly.xlsm", "subfolder": "CSB",
     "purpose": "Crime Suppression Bureau monthly metrics", "tier": 1},
    {"id": 2, "name": "STACP.xlsm", "subfolder": "STACP",
     "purpose": "Strategic Tackling Armed Crime Plan data", "tier": 1},
    {"id": 3, "name": "patrol_monthly.xlsm", "subfolder": "Patrol",
     "purpose": "Patrol Division monthly activity metrics", "tier": 1},
    {"id": 4, "name": "Policy_Training_Monthly.xlsx", "subfolder": "Policy_Training",
     "purpose": "Training course attendance and cost tracking", "tier": 1},
    {"id": 5, "name": "detectives_monthly.xlsx", "subfolder": "Detectives",
     "purpose": "Detective case dispositions and clearance rates", "tier": 1},
    {"id": 6, "name": "ESU.xlsx", "subfolder": "ESU",
     "purpose": "Emergency Services Unit monthly activity", "tier": 1},
    {"id": 7, "name": "Traffic_Monthly.xlsx", "subfolder": "Traffic",
     "purpose": "Traffic Bureau enforcement and citations", "tier": 1},
    {"id": 8, "name": "chief_monthly.xlsx", "subfolder": "Chief",
     "purpose": "Chief's Projects and Initiatives tracking", "tier": 1},
    {"id": 9, "name": "Drone_Monthly.xlsx", "subfolder": "",
     "purpose": "Drone operations metrics (DFR vs Non-DFR)", "tier": 1},
    {"id": 10, "name": "dfr_directed_patrol_enforcement.xlsx", "subfolder": "Drone",
     "purpose": "DFR Summons Log (Claude in Excel workbook; ETL-fed via dfr_export.py)", "tier": 1},
]

TIER2_WORKBOOKS = [
    {"id": 11, "name": "policy_training_outputs.xlsx",
     "path": ONEDRIVE_ROOT / "02_ETL_Scripts" / "Policy_Training_Monthly" / "output" / "policy_training_outputs.xlsx",
     "purpose": "Aggregated training delivery costs (ETL output)", "tier": 2},
    {"id": 12, "name": "Assignment_Master_V3_FINAL.xlsx",
     "path": ONEDRIVE_ROOT / "09_Reference" / "Personnel" / "Assignment_Master_V3_FINAL.xlsx",
     "purpose": "Personnel roster — 25 columns, 166 records", "tier": 2},
    {"id": 13, "name": "NIBRS_Monthly_Report.xlsx",
     "path": ONEDRIVE_ROOT / "01_DataSources" / "NIBRS" / "NIBRS_Monthly_Report.xlsx",
     "purpose": "NIBRS monthly incident classification", "tier": 2},
    {"id": 14, "name": "_SSOCC - Service Log.xlsx",
     "path": ONEDRIVE_ROOT / "KB_Shared" / "04_output" / "ssocc_claude_in_excel_rework" / "_SSOCC - Service Log.xlsx",
     "purpose": "SSOCC dispatch and alert data", "tier": 2},
]

# ---------------------------------------------------------------------------
# Known M-code -> workbook mappings
# ---------------------------------------------------------------------------
M_CODE_MAP = {
    "csb_monthly.xlsm": {
        "m_code": ["m_code/csb/___CSB_Monthly.m"],
        "sheet_target": "MoM sheet",
        "month_format": "MM-YY",
        "pages": ["CSB"],
    },
    "STACP.xlsm": {
        "m_code": ["m_code/stacp/___STACP_pt_1_2.m"],
        "sheet_target": "Multiple sheets (month columns MM-YY)",
        "month_format": "MM-YY",
        "pages": ["STACP"],
    },
    "patrol_monthly.xlsm": {
        "m_code": ["m_code/patrol/___Patrol.m"],
        "sheet_target": "_mom_patrol table",
        "month_format": "MM-YY",
        "pages": ["Patrol"],
    },
    "Policy_Training_Monthly.xlsx": {
        "m_code": ["m_code/training/___In_Person_Training.m"],
        "sheet_target": "Training_Log sheet (fallback Training_Log_Clean)",
        "month_format": "Date column (Start date)",
        "pages": ["Policy & Training Qual"],
    },
    "detectives_monthly.xlsx": {
        "m_code": ["m_code/detectives/___Detectives.m", "m_code/detectives/___Det_case_dispositions_clearance.m"],
        "sheet_target": "Case data tables / Disposition tables",
        "month_format": "MM-YY",
        "pages": ["Detectives"],
    },
    "ESU.xlsx": {
        "m_code": ["m_code/esu/ESU_13Month.m"],
        "sheet_target": "_YY_MMM named tables + _mom_hacsoc dimension",
        "month_format": "_YY_MMM (e.g., _25_JAN)",
        "pages": ["ESU"],
    },
    "Traffic_Monthly.xlsx": {
        "m_code": ["m_code/traffic/___Traffic.m"],
        "sheet_target": "_mom_traffic table",
        "month_format": "MM-YY",
        "pages": ["Traffic"],
    },
    "chief_monthly.xlsx": {
        "m_code": ["m_code/chief/___Chief2.m", "m_code/chief/___chief_projects.m"],
        "sheet_target": "Table8 (events) / Projects table",
        "month_format": "Date column",
        "pages": ["Chief"],
    },
    "Drone_Monthly.xlsx": {
        "m_code": ["m_code/drone/___Drone.m"],
        "sheet_target": "DFR Activity sheet, Non-DFR sheet",
        "month_format": "Date column",
        "pages": ["Drone"],
    },
    "dfr_directed_patrol_enforcement.xlsx": {
        "m_code": ["m_code/drone/DFR_Summons.m"],
        "sheet_target": "DFR_Summons table (fallback: DFR Summons Log sheet)",
        "month_format": "Date column + MM-YY derived",
        "pages": ["DFR Summons"],
        "python_etl": "scripts/dfr_export.py",
    },
    "policy_training_outputs.xlsx": {
        "m_code": ["m_code/training/___In_Person_Training.m", "m_code/training/___Cost_of_Training.m"],
        "sheet_target": "InPerson_Prior_Month_List / Delivery_Cost_By_Month",
        "month_format": "Date column / MM-YY",
        "pages": ["Policy & Training Qual"],
    },
    "NIBRS_Monthly_Report.xlsx": {
        "m_code": ["m_code/nibrs/___NIBRS_Monthly_Report.m"],
        "sheet_target": "NIBRS data tables",
        "month_format": "MM-YY",
        "pages": ["NIBRS"],
    },
    "_SSOCC - Service Log.xlsx": {
        "m_code": ["m_code/ssocc/___SSOCC_Data.m"],
        "sheet_target": "Service log tables",
        "month_format": "Date column",
        "pages": ["SSOCC"],
    },
    "Assignment_Master_V3_FINAL.xlsx": {
        "m_code": [],
        "sheet_target": "Personnel roster",
        "month_format": "N/A",
        "pages": ["(Reference -- not directly visualized)"],
    },
}

RELATED_DOCS = {
    "csb_monthly.xlsm": [],
    "STACP.xlsm": ["docs/stacp_standardization_prompt_claude_in_excel.md"],
    "patrol_monthly.xlsm": [],
    "Policy_Training_Monthly.xlsx": ["docs/POLICY_TRAINING_AUTOMATION_AND_COST_VISUAL.md", "docs/POWER_BI_YTD_MEASURES_AND_PAGE_INSTRUCTIONS.md"],
    "detectives_monthly.xlsx": [],
    "ESU.xlsx": ["docs/ESU_POWER_BI_LOAD_AND_PUBLISH.md"],
    "Traffic_Monthly.xlsx": [],
    "chief_monthly.xlsx": [],
    "Drone_Monthly.xlsx": [],
    "dfr_directed_patrol_enforcement.xlsx": ["docs/DFR_Summons_Claude_Excel_Development_Log.md", "docs/PROMPT_Claude_In_Excel_DFR_Directed_Patrol_Summons_MCode.md"],
    "policy_training_outputs.xlsx": ["docs/POLICY_TRAINING_AUTOMATION_AND_COST_VISUAL.md"],
    "Assignment_Master_V3_FINAL.xlsx": ["09_Reference/Personnel/Assignment_Master_SCHEMA.md"],
    "NIBRS_Monthly_Report.xlsx": [],
    "_SSOCC - Service Log.xlsx": ["docs/SSOCC_Service_Log_Excel_And_Power_BI_Rework_2026_03.md"],
}


# ---------------------------------------------------------------------------
# Config loaders
# ---------------------------------------------------------------------------
def load_scripts_json():
    p = REPO_ROOT / "config" / "scripts.json"
    if p.exists():
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_visual_export_mapping():
    p = REPO_ROOT / "Standards" / "config" / "powerbi_visuals" / "visual_export_mapping.json"
    if p.exists():
        with open(p, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def find_etl_scripts_for_workbook(wb_name, scripts_json):
    results = []
    keywords_map = {
        "csb_monthly.xlsm": ["csb"],
        "STACP.xlsm": ["stacp"],
        "patrol_monthly.xlsm": ["patrol"],
        "Policy_Training_Monthly.xlsx": ["policy", "training"],
        "detectives_monthly.xlsx": ["detective"],
        "ESU.xlsx": ["esu"],
        "Traffic_Monthly.xlsx": ["traffic"],
        "chief_monthly.xlsx": ["chief"],
        "Drone_Monthly.xlsx": ["drone"],
        "dfr_directed_patrol_enforcement.xlsx": ["summon", "dfr"],
        "policy_training_outputs.xlsx": ["policy", "training"],
        "NIBRS_Monthly_Report.xlsx": ["nibrs"],
        "_SSOCC - Service Log.xlsx": ["ssocc"],
        "Assignment_Master_V3_FINAL.xlsx": ["personnel", "assignment"],
    }
    kws = keywords_map.get(wb_name, [])
    for script in scripts_json.get("scripts", []):
        script_kws = [k.lower() for k in script.get("keywords", [])]
        if any(k in script_kws for k in kws):
            results.append(script)
    if wb_name == "dfr_directed_patrol_enforcement.xlsx":
        results.append({"name": "DFR Export", "script": "scripts/dfr_export.py", "enabled": True})
    return results


def find_visuals_for_workbook(wb_name, visual_mapping):
    domain_map = {
        "csb_monthly.xlsm": ["csb"],
        "STACP.xlsm": ["stacp"],
        "patrol_monthly.xlsm": ["patrol"],
        "Policy_Training_Monthly.xlsx": ["training", "policy"],
        "detectives_monthly.xlsx": ["detective"],
        "ESU.xlsx": ["esu"],
        "Traffic_Monthly.xlsx": ["traffic"],
        "chief_monthly.xlsx": ["chief"],
        "Drone_Monthly.xlsx": ["drone"],
        "dfr_directed_patrol_enforcement.xlsx": ["dfr", "drone"],
        "policy_training_outputs.xlsx": ["training"],
        "NIBRS_Monthly_Report.xlsx": ["nibrs"],
        "_SSOCC - Service Log.xlsx": ["ssocc"],
        "Assignment_Master_V3_FINAL.xlsx": [],
    }
    domains = domain_map.get(wb_name, [])
    results = []
    for m in visual_mapping.get("mappings", []):
        folder = m.get("normalized_folder", "").lower()
        page = m.get("page_name", "").lower()
        if any(d in folder or d in page for d in domains):
            results.append(m)
    return results


# ---------------------------------------------------------------------------
# Workbook analysis (READ-ONLY -- never writes to target)
# ---------------------------------------------------------------------------
def analyze_workbook(wb_path):
    """Analyze workbook using read_only mode. Returns info dict."""
    info = {
        "sheets": [],
        "tables": [],
        "columns": {},
        "named_ranges": [],
        "has_vba": wb_path.suffix.lower() == ".xlsm",
        "protected_sheets": [],
    }
    is_xlsm = wb_path.suffix.lower() == ".xlsm"

    # Full load needed for tables/protection (read_only skips those).
    # BUT we never save this workbook -- we only read from it.
    wb = openpyxl.load_workbook(str(wb_path), keep_vba=is_xlsm,
                                data_only=False, read_only=False)
    try:
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            visibility = "Visible"
            if ws.sheet_state == "hidden":
                visibility = "Hidden"
            elif ws.sheet_state == "veryHidden":
                visibility = "Very Hidden"

            table_names = []
            for tbl in ws.tables.values():
                table_names.append(tbl.name)
                info["tables"].append({
                    "name": tbl.name, "sheet": ws_name, "ref": tbl.ref,
                })

            row_count = ws.max_row or 0
            col_count = ws.max_column or 0

            if ws.protection.sheet:
                info["protected_sheets"].append(ws_name)

            info["sheets"].append({
                "name": ws_name, "visibility": visibility,
                "tables": table_names, "rows": row_count, "cols": col_count,
            })

            if visibility == "Visible" and row_count > 0 and ws_name != "AI_Context_Reference":
                headers = []
                for col_idx in range(1, min(col_count + 1, 51)):
                    cell = ws.cell(row=1, column=col_idx)
                    if cell.value is not None:
                        hdr = str(cell.value)
                        headers.append({
                            "name": hdr,
                            "dtype": _infer_dtype(ws, col_idx, row_count),
                            "sample": _get_sample(ws, col_idx),
                            "has_nbsp": "\xa0" in hdr,
                            "validation": "",
                            "dropdown_source": "",
                        })
                if headers:
                    info["columns"][ws_name] = headers

        try:
            for name in wb.defined_names:
                info["named_ranges"].append(name)
        except Exception:
            pass
    finally:
        wb.close()

    return info


def _infer_dtype(ws, col_idx, max_row):
    types_seen = set()
    for r in range(2, min(max_row + 1, 12)):
        cell = ws.cell(row=r, column=col_idx)
        v = cell.value
        if v is None:
            continue
        if isinstance(v, datetime):
            types_seen.add("Date")
        elif isinstance(v, bool):
            types_seen.add("Boolean")
        elif isinstance(v, int):
            types_seen.add("Integer")
        elif isinstance(v, float):
            types_seen.add("Decimal")
        elif isinstance(v, str):
            if re.match(r"^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}$", v):
                types_seen.add("Date")
            elif re.match(r"^\d{1,2}-\d{2}$", v):
                types_seen.add("Text (MM-YY)")
            else:
                types_seen.add("Text")
        if cell.data_type == "f":
            types_seen.add("Formula")
    if not types_seen:
        return "Empty"
    if "Formula" in types_seen:
        return "Formula/Calculated"
    if len(types_seen) == 1:
        return types_seen.pop()
    return " / ".join(sorted(types_seen))


def _get_sample(ws, col_idx):
    for r in range(2, min((ws.max_row or 1) + 1, 7)):
        v = ws.cell(row=r, column=col_idx).value
        if v is not None:
            s = str(v)
            return s[:60]
    return ""


# ---------------------------------------------------------------------------
# Build AI_Context_Reference sheet in a TEMP workbook via openpyxl
# ---------------------------------------------------------------------------
FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=NAVY)
FONT_SUBTITLE = Font(name="Calibri", size=10, italic=True, color=GRAY_TEXT)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=WHITE)
FONT_COL_HDR = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_DATA = Font(name="Calibri", size=10, color=BLACK)
FONT_KV_LABEL = Font(name="Calibri", size=10, bold=True, color=NAVY)
FONT_SUBSECTION = Font(name="Calibri", size=10, bold=True, italic=True, color=NAVY)

FILL_TITLE = PatternFill(start_color=LIGHT_GOLD, end_color=LIGHT_GOLD, fill_type="solid")
FILL_SECTION = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
FILL_COL_HDR = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
FILL_ALT = PatternFill(start_color=GRAY_ALT, end_color=GRAY_ALT, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
BORDER_BOTTOM = Border(bottom=Side(style="thin"))
ALIGN_WRAP = Alignment(wrap_text=True, vertical="top")
ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _write_title(ws, row, text, mc=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_TITLE; c.fill = FILL_TITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 30
    return row + 1


def _write_subtitle(ws, row, text, mc=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_SUBTITLE
    c.alignment = Alignment(horizontal="left", vertical="center")
    return row + 1


def _write_section(ws, row, text, mc=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=mc)
    c = ws.cell(row=row, column=1, value=text)
    c.font = FONT_SECTION; c.fill = FILL_SECTION
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def _write_col_hdrs(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = FONT_COL_HDR; c.fill = FILL_COL_HDR
        c.border = BORDER_BOTTOM; c.alignment = ALIGN_WRAP
    return row + 1


def _write_kv(ws, row, label, value, alt=False):
    c1 = ws.cell(row=row, column=1, value=label)
    c1.font = FONT_KV_LABEL; c1.alignment = ALIGN_LEFT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    c2 = ws.cell(row=row, column=2, value=value)
    c2.font = FONT_DATA; c2.alignment = ALIGN_LEFT
    if alt:
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = FILL_ALT
    return row + 1


def _write_data(ws, row, values, alt=False):
    for i, v in enumerate(values, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.font = FONT_DATA; c.alignment = ALIGN_WRAP
        c.fill = FILL_ALT if alt else FILL_WHITE
    return row + 1


def _auto_fit(ws, min_w=15, max_w=80):
    for ci in range(1, (ws.max_column or 1) + 1):
        ml = min_w
        cl = get_column_letter(ci)
        for ri in range(1, min((ws.max_row or 1) + 1, 200)):
            v = ws.cell(row=ri, column=ci).value
            if v:
                ml = max(ml, min(len(str(v)), max_w))
        ws.column_dimensions[cl].width = min(ml + 2, max_w)


def build_temp_workbook(wb_path, wb_entry, info, scripts_json, visual_mapping):
    """Build a temp .xlsx containing only the AI_Context_Reference sheet.
    Returns path to temp file."""
    mc = M_CODE_MAP.get(wb_entry["name"], {})
    m_code_files = mc.get("m_code", [])
    is_xlsm = wb_path.suffix.lower() == ".xlsm"

    etl_scripts = find_etl_scripts_for_workbook(wb_entry["name"], scripts_json)
    python_etl = mc.get("python_etl", "")
    if not python_etl and etl_scripts:
        python_etl = "; ".join(f"{s['name']} ({s.get('script','')})" for s in etl_scripts[:3])
    if not python_etl:
        python_etl = "None -- direct Power Query load"

    visuals = find_visuals_for_workbook(wb_entry["name"], visual_mapping)
    visual_names = [v["visual_name"] for v in visuals]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI_Context_Reference"
    ws.sheet_properties.tabColor = GOLD
    MC = 7

    row = 1
    row = _write_title(ws, row, f"{wb_entry['name']} -- AI & ETL Context Architecture", MC)
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    row = _write_subtitle(ws, row, f"Generated: {ts} | Hackensack PD Compstat Pipeline", MC)
    row += 1

    # Section 1: OVERVIEW
    row = _write_section(ws, row, "1. OVERVIEW", MC)
    tier_label = f"Tier {wb_entry['tier']}"
    role = f"{tier_label} -- {'Staff Data Entry -> Power BI via M Code' if wb_entry['tier'] == 1 else 'ETL Output / Reference'}"
    for i, (l, v) in enumerate([
        ("Purpose", wb_entry["purpose"]),
        ("File Format", wb_path.suffix + (" (macro-enabled)" if is_xlsm else "")),
        ("Pipeline Role", role),
        ("OneDrive Path", str(wb_path)),
        ("Consuming M Code", ", ".join(m_code_files) if m_code_files else "None"),
        ("Python ETL", python_etl),
        ("Power BI Page(s)", ", ".join(mc.get("pages", ["(Unknown)"]))),
        ("13-Month Window", "Yes -- pReportMonth driven | Standard pattern" if m_code_files else "N/A"),
    ]):
        row = _write_kv(ws, row, l, v, alt=(i % 2 == 1))
    row += 1

    # Section 2: SHEET & TABLE DIRECTORY
    row = _write_section(ws, row, "2. SHEET & TABLE DIRECTORY", MC)
    row = _write_col_hdrs(ws, row, ["Sheet Name", "Visibility", "Table Name(s)", "Row Count", "Column Count", "Notes", ""])
    for i, sh in enumerate(info["sheets"]):
        notes = ""
        if sh["name"] in info["protected_sheets"]:
            notes = "Protected"
        if any(c.get("has_nbsp") for c in info["columns"].get(sh["name"], [])):
            notes += (" | " if notes else "") + "Has \\xa0 in headers"
        tbl_str = ", ".join(sh["tables"]) if sh["tables"] else "--"
        row = _write_data(ws, row, [sh["name"], sh["visibility"], tbl_str, sh["rows"], sh["cols"], notes, ""], alt=(i % 2 == 1))
    row += 1

    # Section 3: DATA DICTIONARY
    row = _write_section(ws, row, "3. DATA DICTIONARY & VALIDATION", MC)
    row = _write_col_hdrs(ws, row, ["Column Name", "Data Type", "Sample Value", "Validation/Dropdown", "Dropdown Source", "Formula Dependencies", "Notes"])
    dri = 0
    for ws_name, cols in info["columns"].items():
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MC)
        c = ws.cell(row=row, column=1, value=f"-- {ws_name} --")
        c.font = FONT_SUBSECTION
        row += 1
        for ci in cols:
            n = "Contains non-breaking space (\\xa0)" if ci.get("has_nbsp") else ""
            row = _write_data(ws, row, [ci["name"], ci["dtype"], ci["sample"], ci.get("validation", ""), ci.get("dropdown_source", ""), "", n], alt=(dri % 2 == 1))
            dri += 1
    row += 1

    # Section 4: ETL & INTEGRATION MAP
    row = _write_section(ws, row, "4. ETL & INTEGRATION MAP", MC)
    for i, (l, v) in enumerate([
        ("M Code Query", ", ".join(m_code_files) if m_code_files else "None"),
        ("M Code Source Pattern", f'Excel.Workbook(File.Contents("{wb_path}"), null, true)' if m_code_files else "N/A"),
        ("Target Sheet/Table", mc.get("sheet_target", "Unknown")),
        ("Month Column Format", mc.get("month_format", "Unknown")),
        ("Rolling Window", "pReportMonth -> EndOfMonth(pReportMonth) back 12 months" if m_code_files else "N/A"),
        ("Python ETL Script", python_etl),
        ("Output CSV Pattern", "; ".join(s.get("output_patterns", ["N/A"])[0] for s in etl_scripts[:3]) if etl_scripts else "N/A"),
        ("Downstream Power BI Visuals", "; ".join(visual_names[:5]) if visual_names else "None mapped"),
    ]):
        row = _write_kv(ws, row, l, v, alt=(i % 2 == 1))
    row += 1

    # Section 5: VBA (xlsm only)
    if is_xlsm:
        row = _write_section(ws, row, "5. VBA & MACROS", MC)
        row = _write_col_hdrs(ws, row, ["Module Name", "Type", "Purpose", "Triggers", "", "", ""])
        row = _write_data(ws, row, ["(VBA project present)", "Standard/Class/UserForm", "Inspection requires xlwings or COM automation", "--", "", "", ""])
        row += 1

    # Section 6: CLAUDE IN EXCEL QUICK START
    row = _write_section(ws, row, "6. CLAUDE IN EXCEL -- QUICK START", MC)
    table_names = [t["name"] for t in info["tables"]]
    gotchas = []
    for wn, cols in info["columns"].items():
        nbsp = [c["name"] for c in cols if c.get("has_nbsp")]
        if nbsp:
            gotchas.append(f"Non-breaking spaces in headers on '{wn}': {nbsp}")
    if info["protected_sheets"]:
        gotchas.append(f"Protected sheets: {info['protected_sheets']}")
    if is_xlsm:
        gotchas.append("Macro-enabled workbook -- use keep_vba=True when saving with openpyxl")
    related = RELATED_DOCS.get(wb_entry["name"], [])
    for i, (l, v) in enumerate([
        ("Workbook Purpose", wb_entry["purpose"]),
        ("Key Tables", ", ".join(table_names) if table_names else "No structured tables found"),
        ("Month Key Format", mc.get("month_format", "Unknown")),
        ("Common Formulas", "Check individual sheets for XLOOKUP / VLOOKUP / INDEX-MATCH patterns"),
        ("Protected Ranges", ", ".join(info["protected_sheets"]) if info["protected_sheets"] else "None detected"),
        ("Validation Rules", "See Data Dictionary section above for per-column details"),
        ("Gotchas", "; ".join(gotchas) if gotchas else "None detected"),
        ("Related Docs", "; ".join(related) if related else "See docs/ in 06_Workspace_Management"),
    ]):
        row = _write_kv(ws, row, l, v, alt=(i % 2 == 1))

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    _auto_fit(ws)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    wb.save(tmp.name)
    wb.close()
    return tmp.name


# ---------------------------------------------------------------------------
# ZIP-LEVEL INJECTION v3 -- TEXT-LEVEL XML manipulation (no ElementTree on target)
# ---------------------------------------------------------------------------
# Strategy: NEVER parse target XML through ElementTree. Instead, read the raw
# bytes of workbook.xml, workbook.xml.rels, and [Content_Types].xml, then use
# regex/string insertion to add our entries. This preserves 100% of the original
# XML including namespace declarations, mc:Ignorable, extension elements, etc.
# We also skip styles.xml entirely -- all cells use s="0" (default style).
# The injected sheet won't have HPD formatting, but the workbook stays intact.

NS_SS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


def _build_plain_sheet_xml(temp_wb_path):
    """Extract sheet data from temp workbook and build a plain sheet XML
    with inline strings and no style references (s='0' default)."""
    # Read shared strings and sheet XML from temp workbook
    with zipfile.ZipFile(temp_wb_path, "r") as tz:
        sheet_bytes = tz.read("xl/worksheets/sheet1.xml")
        shared_strings = []
        if "xl/sharedStrings.xml" in tz.namelist():
            ss_tree = ET.parse(io.BytesIO(tz.read("xl/sharedStrings.xml")))
            for si in ss_tree.getroot().findall(f"{{{NS_SS}}}si"):
                # Handle both simple <t> and rich text <r><t> elements
                parts = []
                t_el = si.find(f"{{{NS_SS}}}t")
                if t_el is not None and t_el.text:
                    parts.append(t_el.text)
                else:
                    for r_el in si.findall(f"{{{NS_SS}}}r"):
                        rt = r_el.find(f"{{{NS_SS}}}t")
                        if rt is not None and rt.text:
                            parts.append(rt.text)
                shared_strings.append("".join(parts))

    # Parse sheet XML with ElementTree (this is the TEMP file, not the target)
    ET.register_namespace("", NS_SS)
    ET.register_namespace("r", "http://schemas.openxmlformats.org/officeDocument/2006/relationships")
    sheet_tree = ET.ElementTree(ET.fromstring(sheet_bytes))
    sheet_root = sheet_tree.getroot()

    # Convert shared-string refs to inline strings
    for c_el in sheet_root.iter(f"{{{NS_SS}}}c"):
        if c_el.get("t") == "s":
            v_el = c_el.find(f"{{{NS_SS}}}v")
            if v_el is not None and v_el.text is not None:
                idx = int(v_el.text)
                text_val = shared_strings[idx] if idx < len(shared_strings) else ""
                c_el.set("t", "inlineStr")
                c_el.remove(v_el)
                is_el = ET.SubElement(c_el, f"{{{NS_SS}}}is")
                t_el = ET.SubElement(is_el, f"{{{NS_SS}}}t")
                t_el.text = text_val
                if text_val and (text_val[0] == " " or text_val[-1] == " "):
                    t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")

        # Strip style index -- use default style
        if c_el.get("s") is not None:
            del c_el.attrib["s"]

    # Remove merge cells, sheet views, etc. that might reference styles
    # Keep it simple -- just sheetData + mergeCells
    buf = io.BytesIO()
    sheet_tree.write(buf, xml_declaration=True, encoding="UTF-8")
    return buf.getvalue()


def _escape_xml(text):
    """Escape text for safe XML attribute/content insertion."""
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")


def inject_via_zip(target_path, temp_wb_path, sheet_name="AI_Context_Reference"):
    """Inject sheet using TEXT-LEVEL XML manipulation on target files.

    Key principle: we NEVER parse target XML through ElementTree.
    We read raw bytes, do regex-based insertions, and write back.
    Only the temp workbook's sheet XML goes through ElementTree (safe).
    """
    # Build the sheet XML from temp workbook (ET only touches temp file)
    new_sheet_xml = _build_plain_sheet_xml(temp_wb_path)

    # --- Read raw bytes from target zip ---
    with zipfile.ZipFile(str(target_path), "r") as zf:
        target_names = zf.namelist()
        wb_xml = zf.read("xl/workbook.xml").decode("utf-8")
        rels_xml = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")
        ct_xml = zf.read("[Content_Types].xml").decode("utf-8")

    # --- Determine next sheet number ---
    sheet_nums = set()
    for fn in target_names:
        m = re.search(r"xl/worksheets/sheet(\d+)\.xml$", fn.replace("\\", "/"))
        if m:
            sheet_nums.add(int(m.group(1)))
    new_sheet_num = max(sheet_nums, default=0) + 1
    new_sheet_file = f"xl/worksheets/sheet{new_sheet_num}.xml"
    new_sheet_rel_target = f"worksheets/sheet{new_sheet_num}.xml"

    # --- Determine next rId ---
    rid_nums = set()
    for m in re.finditer(r'Id="rId(\d+)"', rels_xml):
        rid_nums.add(int(m.group(1)))
    new_rid = f"rId{max(rid_nums, default=0) + 1}"

    # --- Determine next sheetId ---
    sheet_ids = set()
    for m in re.finditer(r'sheetId="(\d+)"', wb_xml):
        sheet_ids.add(int(m.group(1)))
    new_sheet_id = max(sheet_ids, default=0) + 1

    # --- Check if sheet already exists (idempotent) ---
    old_sheet_pattern = re.compile(
        r'<sheet[^>]*name="' + re.escape(sheet_name) + r'"[^>]*/>', re.DOTALL
    )
    old_match = old_sheet_pattern.search(wb_xml)
    old_rid = None
    old_sheet_file = None

    if old_match:
        # Extract old rId
        rid_match = re.search(r'r:id="(rId\d+)"', old_match.group(0))
        if rid_match:
            old_rid = rid_match.group(1)
        # Remove old sheet element from workbook.xml
        wb_xml = old_sheet_pattern.sub("", wb_xml)

    if old_rid:
        # Find and remove old relationship
        old_rel_pattern = re.compile(
            r'<Relationship[^>]*Id="' + re.escape(old_rid) + r'"[^>]*/>', re.DOTALL
        )
        old_rel_match = old_rel_pattern.search(rels_xml)
        if old_rel_match:
            # Extract old target file
            target_match = re.search(r'Target="([^"]*)"', old_rel_match.group(0))
            if target_match:
                old_sheet_file = f"xl/{target_match.group(1)}"
            rels_xml = old_rel_pattern.sub("", rels_xml)

        # Remove old content type override
        if old_sheet_file:
            old_ct_pattern = re.compile(
                r'<Override[^>]*PartName="/' + re.escape(old_sheet_file) + r'"[^>]*/>', re.DOTALL
            )
            ct_xml = old_ct_pattern.sub("", ct_xml)

    # --- Insert new sheet entry into workbook.xml ---
    sheet_entry = (
        f'<sheet name="{_escape_xml(sheet_name)}" sheetId="{new_sheet_id}" '
        f'r:id="{new_rid}"/>'
    )
    # Insert before </sheets>
    wb_xml = wb_xml.replace("</sheets>", f"{sheet_entry}</sheets>")

    # --- Insert new relationship into workbook.xml.rels ---
    rel_entry = (
        f'<Relationship Id="{new_rid}" '
        f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" '
        f'Target="{new_sheet_rel_target}"/>'
    )
    wb_xml_rels_close = "</Relationships>"
    rels_xml = rels_xml.replace(wb_xml_rels_close, f"{rel_entry}{wb_xml_rels_close}")

    # --- Insert new content type override into [Content_Types].xml ---
    ct_entry = (
        f'<Override PartName="/{new_sheet_file}" '
        f'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>'
    )
    ct_xml = ct_xml.replace("</Types>", f"{ct_entry}</Types>")

    # --- Rewrite the zip ---
    skip_files = set()
    if old_sheet_file:
        skip_files.add(old_sheet_file)
    # We ONLY replace the 3 metadata files + add the new sheet.
    # styles.xml and all other files are copied byte-for-byte.
    replace_files = {
        "xl/workbook.xml": wb_xml.encode("utf-8"),
        "xl/_rels/workbook.xml.rels": rels_xml.encode("utf-8"),
        "[Content_Types].xml": ct_xml.encode("utf-8"),
    }

    tmp_out = tempfile.NamedTemporaryFile(suffix=target_path.suffix, delete=False,
                                          dir=str(target_path.parent))
    tmp_out.close()

    with zipfile.ZipFile(str(target_path), "r") as src_zip:
        with zipfile.ZipFile(tmp_out.name, "w", zipfile.ZIP_DEFLATED) as dst_zip:
            for item in src_zip.infolist():
                if item.filename in skip_files:
                    continue
                if item.filename in replace_files:
                    dst_zip.writestr(item, replace_files[item.filename])
                else:
                    # Byte-for-byte copy -- preserves everything
                    dst_zip.writestr(item, src_zip.read(item.filename))

            # Add the new sheet
            dst_zip.writestr(new_sheet_file, new_sheet_xml)

    # Replace original with new
    os.replace(tmp_out.name, str(target_path))


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------
def backup_workbook(wb_path):
    """Create a timestamped backup. Returns backup path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak_dir = wb_path.parent / "archive"
    bak_dir.mkdir(exist_ok=True)
    bak_name = f"{wb_path.stem}_{ts}_pre_injection{wb_path.suffix}"
    bak_path = bak_dir / bak_name
    shutil.copy2(str(wb_path), str(bak_path))
    return bak_path


# ---------------------------------------------------------------------------
# Main injection orchestrator
# ---------------------------------------------------------------------------
def inject_ai_context_sheet(wb_path, wb_entry, scripts_json, visual_mapping,
                            dry_run=False, verbose=False, do_backup=True):
    result = {
        "workbook": wb_entry["name"],
        "status": "X Error",
        "sheets": 0, "tables": 0,
        "vba_modules": "N/A", "m_code": "", "reason": "",
    }

    if not wb_path.exists():
        result["status"] = "!! Skipped"
        result["reason"] = f"File not found: {wb_path}"
        return result

    # Analyze (read-only -- never writes to target)
    try:
        info = analyze_workbook(wb_path)
    except Exception as e:
        result["status"] = "X Error"
        result["reason"] = f"Analysis failed: {e}"
        return result

    result["sheets"] = len(info["sheets"])
    result["tables"] = len(info["tables"])
    if info["has_vba"]:
        result["vba_modules"] = "VBA present"

    mc = M_CODE_MAP.get(wb_entry["name"], {})
    m_code_files = mc.get("m_code", [])
    result["m_code"] = ", ".join(m_code_files) if m_code_files else "None"

    if dry_run:
        result["status"] = ">> Dry Run"
        result["reason"] = "Would inject AI_Context_Reference"
        if verbose:
            print(f"  Sheets: {[s['name'] for s in info['sheets']]}")
            print(f"  Tables: {[t['name'] for t in info['tables']]}")
            print(f"  M Code: {m_code_files}")
            for wn, cols in info["columns"].items():
                nbsp = [c["name"] for c in cols if c.get("has_nbsp")]
                if nbsp:
                    print(f"  !! Non-breaking spaces in '{wn}': {nbsp}")
        return result

    # Backup
    if do_backup:
        try:
            bak = backup_workbook(wb_path)
            if verbose:
                print(f"      Backup: {bak}")
        except Exception as e:
            result["status"] = "X Error"
            result["reason"] = f"Backup failed: {e}"
            return result

    # Build temp workbook with the reference sheet
    try:
        temp_path = build_temp_workbook(wb_path, wb_entry, info, scripts_json, visual_mapping)
    except Exception as e:
        result["status"] = "X Error"
        result["reason"] = f"Build failed: {e}"
        return result

    # Inject via zip
    try:
        inject_via_zip(wb_path, temp_path, "AI_Context_Reference")
        result["status"] = "OK Injected"
    except PermissionError:
        result["status"] = "!! Skipped"
        result["reason"] = "File is open or permission denied"
    except Exception as e:
        result["status"] = "X Error"
        result["reason"] = f"Injection failed: {e}"
    finally:
        try:
            os.unlink(temp_path)
        except OSError:
            pass

    return result


# ---------------------------------------------------------------------------
# Path resolution
# ---------------------------------------------------------------------------
def resolve_path(entry):
    if "path" in entry:
        return Path(entry["path"])
    subfolder = entry.get("subfolder", "")
    if subfolder:
        return CONTRIBUTIONS_DIR / subfolder / entry["name"]
    return CONTRIBUTIONS_DIR / entry["name"]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="Inject AI_Context_Reference into HPD shared workbooks (v2 zip-safe)")
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing")
    parser.add_argument("--tier2", action="store_true", help="Include Tier 2 workbooks")
    parser.add_argument("--workbook", type=str, help="Filter to a single workbook (substring match)")
    parser.add_argument("--verbose", action="store_true", help="Detailed logging")
    parser.add_argument("--no-backup", action="store_true", help="Skip pre-injection backup (not recommended)")
    args = parser.parse_args()

    print("=" * 72)
    print("  AI_Context_Reference Sheet Injector v2 (zip-safe)")
    print(f"  {'DRY RUN MODE' if args.dry_run else 'LIVE MODE'}")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Backup: {'DISABLED' if args.no_backup else 'enabled (archive/ subfolder)'}")
    print("=" * 72)

    scripts_json = load_scripts_json()
    visual_mapping = load_visual_export_mapping()
    if scripts_json:
        print(f"  Loaded scripts.json: {len(scripts_json.get('scripts', []))} scripts")
    if visual_mapping:
        print(f"  Loaded visual_export_mapping.json: {len(visual_mapping.get('mappings', []))} mappings")
    print()

    workbooks = list(TIER1_WORKBOOKS)
    if args.tier2:
        workbooks.extend(TIER2_WORKBOOKS)

    if args.workbook:
        pattern = args.workbook.lower()
        workbooks = [w for w in workbooks if pattern in w["name"].lower()]
        if not workbooks:
            print(f"  No workbooks matched filter '{args.workbook}'")
            sys.exit(1)

    print(f"  Target workbooks: {len(workbooks)}")
    print(f"  Contributions dir: {CONTRIBUTIONS_DIR}")
    print()

    results = []
    for entry in workbooks:
        wb_path = resolve_path(entry)
        label = f"[{entry['id']:>2}] {entry['name']}"
        print(f"  Processing {label} ...")
        if args.verbose:
            print(f"      Path: {wb_path}")

        r = inject_ai_context_sheet(wb_path, entry, scripts_json, visual_mapping,
                                    dry_run=args.dry_run, verbose=args.verbose,
                                    do_backup=not args.no_backup)
        results.append(r)
        status = r["status"]
        print(f"      -> {status}" + (f" ({r['reason']})" if r["reason"] else ""))
        print()

    # Summary
    print()
    print("=" * 110)
    print(f"  {'Workbook':<42} {'Status':<16} {'Sheets':>6} {'Tables':>7} {'VBA':>10} {'M Code':<30}")
    print("-" * 110)
    for r in results:
        print(f"  {r['workbook']:<42} {r['status']:<16} {r['sheets']:>6} {r['tables']:>7} {str(r['vba_modules'])[:10]:>10} {r['m_code'][:30]:<30}")
    print("=" * 110)

    ok = sum(1 for r in results if "Injected" in r["status"] or "Dry Run" in r["status"])
    skip = sum(1 for r in results if "Skipped" in r["status"])
    err = sum(1 for r in results if "Error" in r["status"])
    print(f"\n  Total: {len(results)} | Success: {ok} | Skipped: {skip} | Errors: {err}")

    if any(r["reason"] for r in results):
        print("\n  Notes:")
        for r in results:
            if r["reason"]:
                print(f"    {r['workbook']}: {r['reason']}")


if __name__ == "__main__":
    main()
