"""
dfr_backfill_descriptions.py
Backfill Description, Fine Amount, Source Type, Violation Type in DFR Summons Log.

Problem: XLOOKUP formulas in the DFR workbook normalize statute codes by stripping
parentheses only (e.g., 88-6D(2) -> 88-6D2), but ViolationData only has parent codes
(e.g., 88-6). This causes most lookups to return empty strings.

Fix: This script uses cascading normalization (exact -> strip parens -> strip to parent)
to resolve statutes against ViolationData, then writes values directly into the
DFR Summons Log cells so Power BI reads cached values.

Also adds missing subsection codes to ViolationData sheet for future XLOOKUP success.

Usage:
    python dfr_backfill_descriptions.py           # Dry run (report only)
    python dfr_backfill_descriptions.py --apply    # Write changes to workbook
"""

from __future__ import annotations

import argparse
import logging
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

try:
    from path_config import get_onedrive_root
except ImportError:
    import os

    def get_onedrive_root() -> Path:
        base = os.environ.get("ONEDRIVE_BASE") or os.environ.get("ONEDRIVE_HACKENSACK")
        if base:
            return Path(base)
        return Path.home() / "OneDrive - City of Hackensack"


logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")

# Columns in DFR Summons Log that need backfill (0-indexed from col A)
COL_MAP = {
    "Statute": 6,          # F (1-indexed) = index 5, but openpyxl is 1-indexed
    "Description": 7,      # G
    "Fine Amount": 8,      # H
    "Source Type": 9,       # I
    "Violation Type": 10,   # J
}


def _normalize_statute(code: str) -> list[str]:
    """Return list of candidate codes for cascading lookup.
    88-6D(2) -> [88-6D(2), 88-6D2, 88-6D, 88-6]
    117-3(F) -> [117-3(F), 117-3F, 117-3]
    39:3-40  -> [39:3-40]
    """
    candidates = [code]
    # Strip parentheses: 88-6D(2) -> 88-6D2
    stripped = code.replace("(", "").replace(")", "")
    if stripped != code:
        candidates.append(stripped)
    # Strip trailing alpha: 88-6D2 -> 88-6D -> 88-6
    s1 = re.sub(r"[0-9]+$", "", stripped).rstrip("-")
    if s1 and s1 not in candidates:
        candidates.append(s1)
    s2 = re.sub(r"[A-Za-z]+$", "", s1).rstrip("-")
    if s2 and s2 not in candidates:
        candidates.append(s2)
    return candidates


def _build_vd_lookup(vd_df: pd.DataFrame) -> dict[str, dict]:
    """Build ViolationData lookup keyed by ViolationCode (case-insensitive)."""
    lookup = {}
    for _, row in vd_df.iterrows():
        code = str(row.get("ViolationCode", "")).strip()
        if code:
            lookup[code.lower()] = {
                "description": str(row.get("Description", "")).strip(),
                "fine_amount": row.get("FineAmount", "0"),
                "source_type": str(row.get("SourceType", "")).strip(),
                "violation_type": str(row.get("ViolationType", "")).strip(),
                "violation_category": str(row.get("ViolationCategory", "")).strip(),
                "normalized_code": str(row.get("NormalizedCode", "")).strip(),
                "display_name": str(row.get("DisplayName", "")).strip(),
            }
    return lookup


def _resolve_statute(statute: str, vd_lookup: dict[str, dict]) -> dict | None:
    """Cascading lookup: exact -> strip parens -> strip trailing alpha -> parent."""
    candidates = _normalize_statute(statute)
    for c in candidates:
        result = vd_lookup.get(c.lower())
        if result and result.get("description"):
            return result
    return None


def backfill(apply: bool = False) -> dict:
    """Run backfill on DFR workbook."""
    base = get_onedrive_root()
    wb_path = base / "Shared Folder" / "Compstat" / "Contributions" / "Drone" / "dfr_directed_patrol_enforcement.xlsx"

    if not wb_path.exists():
        logger.error("DFR workbook not found: %s", wb_path)
        return {"error": "DFR workbook not found"}

    # Load ViolationData via pandas for lookup
    vd_df = pd.read_excel(wb_path, sheet_name="ViolationData", dtype=str)
    vd_lookup = _build_vd_lookup(vd_df)
    logger.info("ViolationData: %d entries loaded", len(vd_lookup))

    # Open workbook with openpyxl to read formulas and write values
    owb = openpyxl.load_workbook(wb_path)
    ws = owb["DFR Summons Log"]

    # Find column indices from header row
    headers = {}
    for cell in ws[1]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    col_statute = headers.get("Statute")
    col_desc = headers.get("Description")
    col_fine = headers.get("Fine Amount")
    col_source = headers.get("Source Type")
    col_vtype = headers.get("Violation Type")
    col_date = headers.get("Date")

    if not all([col_statute, col_desc, col_fine, col_source, col_vtype, col_date]):
        logger.error("Missing required columns. Found: %s", headers)
        return {"error": "Missing columns"}

    logger.info("Column mapping: Statute=%s, Description=%s, Fine=%s, Source=%s, VType=%s",
                col_statute, col_desc, col_fine, col_source, col_vtype)

    filled = 0
    skipped = 0
    already_filled = 0
    unresolved = []
    subsection_codes_to_add = {}

    for row_num in range(2, ws.max_row + 1):
        date_val = ws.cell(row_num, col_date).value
        if date_val is None:
            continue  # Skip empty rows

        statute_val = ws.cell(row_num, col_statute).value
        if not statute_val:
            continue

        statute = str(statute_val).strip()
        desc_cell = ws.cell(row_num, col_desc)
        fine_cell = ws.cell(row_num, col_fine)
        source_cell = ws.cell(row_num, col_source)
        vtype_cell = ws.cell(row_num, col_vtype)

        # Check if Description needs backfill (formula or empty)
        desc_val = desc_cell.value
        is_formula = isinstance(desc_val, str) and desc_val.startswith("=")
        is_empty = desc_val is None or (isinstance(desc_val, str) and desc_val.strip() == "")

        if not is_formula and not is_empty:
            already_filled += 1
            continue

        # Resolve statute
        result = _resolve_statute(statute, vd_lookup)
        if not result:
            if statute not in [u["statute"] for u in unresolved]:
                unresolved.append({"statute": statute, "row": row_num})
            skipped += 1
            continue

        # Track subsection codes that needed parent fallback
        exact = vd_lookup.get(statute.lower())
        if not exact or not exact.get("description"):
            # This statute needed fallback — add to ViolationData
            subsection_codes_to_add[statute] = result

        if apply:
            # Write values (not formulas) into cells
            desc_cell.value = result["description"].upper()
            try:
                fine_cell.value = float(result["fine_amount"]) if result["fine_amount"] else 0
            except (ValueError, TypeError):
                fine_cell.value = 0
            source_cell.value = result["source_type"]
            vtype_cell.value = result["violation_type"]

        filled += 1
        if filled <= 5:
            logger.info("  Row %d: %s -> %s ($%s) [%s]",
                        row_num, statute, result["description"][:50],
                        result["fine_amount"], result["source_type"])

    # --- Add missing subsection codes to ViolationData ---
    if subsection_codes_to_add:
        ws_vd = owb["ViolationData"]
        vd_headers = {}
        for cell in ws_vd[1]:
            if cell.value:
                vd_headers[str(cell.value).strip()] = cell.column

        next_row = ws_vd.max_row + 1
        added_codes = 0
        for code, parent_data in subsection_codes_to_add.items():
            # Check if already exists
            exists = False
            for r in range(2, ws_vd.max_row + 1):
                existing = ws_vd.cell(r, vd_headers.get("ViolationCode", 1)).value
                if existing and str(existing).strip().lower() == code.lower():
                    exists = True
                    break
            if exists:
                continue

            if apply:
                ws_vd.cell(next_row, vd_headers.get("ViolationCode", 1)).value = code
                ws_vd.cell(next_row, vd_headers.get("Description", 2)).value = parent_data["description"]
                ws_vd.cell(next_row, vd_headers.get("FineAmount", 3)).value = float(parent_data.get("fine_amount", 0) or 0)
                ws_vd.cell(next_row, vd_headers.get("SourceType", 4)).value = parent_data["source_type"]
                ws_vd.cell(next_row, vd_headers.get("ViolationType", 5)).value = parent_data["violation_type"]
                ws_vd.cell(next_row, vd_headers.get("ViolationCategory", 6)).value = parent_data.get("violation_category", "")
                # NormalizedCode = lowercase, no parens
                norm = code.replace("(", "").replace(")", "").lower()
                ws_vd.cell(next_row, vd_headers.get("NormalizedCode", 7)).value = norm
                ws_vd.cell(next_row, vd_headers.get("DisplayName", 8)).value = f"{code} - {parent_data['description']}"
                next_row += 1

            added_codes += 1
            logger.info("  ViolationData +%s -> %s (from parent)", code, parent_data["description"][:50])

    # Save if applying
    if apply:
        owb.save(wb_path)
        logger.info("Workbook saved: %s", wb_path)
    else:
        logger.info("DRY RUN — no changes written. Use --apply to write.")

    # --- Report ---
    print("\n" + "=" * 70)
    print("  DFR DESCRIPTION BACKFILL REPORT")
    print("=" * 70)
    print(f"\n  Rows backfilled:         {filled}")
    print(f"  Rows already filled:     {already_filled}")
    print(f"  Rows skipped (no match): {skipped}")
    print(f"  ViolationData codes added: {len(subsection_codes_to_add)}")
    if subsection_codes_to_add:
        for code, data in subsection_codes_to_add.items():
            print(f"    + {code} -> {data['description'][:50]}")
    if unresolved:
        print("\n  Unresolved statutes:")
        for u in unresolved:
            print(f"    {u['statute']} (row {u['row']})")
    if not apply:
        print("\n  *** DRY RUN — rerun with --apply to write changes ***")
    print("\n" + "=" * 70)

    return {
        "filled": filled,
        "already_filled": already_filled,
        "skipped": skipped,
        "subsection_codes_added": len(subsection_codes_to_add),
        "unresolved": unresolved,
    }


def main():
    parser = argparse.ArgumentParser(description="Backfill DFR Description/Fine from ViolationData")
    parser.add_argument("--apply", action="store_true", help="Write changes to workbook (default: dry run)")
    args = parser.parse_args()
    backfill(apply=args.apply)


if __name__ == "__main__":
    main()
