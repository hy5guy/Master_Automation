"""
dfr_export.py
DFR (Drone / Directed Field Report) workbook export module.

Maps DFR-badge summons records from the E-ticket ETL into the
dfr_directed_patrol_enforcement.xlsx log workbook (DFR Summons Log sheet /
DFR_Summons table), deduplicates on Summons Number, and skips formula columns.

Target workbook:
  <OneDrive>/Shared Folder/Compstat/Contributions/Drone/dfr_directed_patrol_enforcement.xlsx

Invoked from run_summons_etl.py after split_dfr_records() isolates DFR rows.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd

logger = logging.getLogger(__name__)

# Column names in the DFR Summons Log that contain Excel formulas.
# These are auto-populated by Excel (XLOOKUP, concatenation, etc.) and must
# never be overwritten by the ETL. The ETL only fills writable data columns.
DFR_FORMULA_COLUMN_NAMES: frozenset[str] = frozenset(
    {
        "Summons ID",       # Col A — auto-generated sequence
        "Description",      # Col G — XLOOKUP from ViolationData
        "Fine Amount",      # Col H — XLOOKUP from ViolationData
        "Violation Type",   # Col I — XLOOKUP P/M/C code
        "DFR Unit ID",      # Col J — formula-derived unit identifier
        "Summons Recall",   # Col P — auto-formula from Summons Status
        "Full Summons Number",  # Col R — "E26" & Summons Number
    }
)

# Map from normalized ETL column names → DFR Summons Log column headers
_ETL_TO_DFR: dict[str, str] = {
    "ISSUE_DATE":           "Date",
    "Charge Time":          "Time",
    "TICKET_NUMBER":        "Summons Number",
    "Offense Street Name":  "Location",
    "STATUTE":              "Statute",
    "OFFICER_DISPLAY_NAME": "DFR Operator",
    # Issuing Officer is the same person for drone enforcement
    "STATUS":               "Summons Status",
}


def _map_to_dfr_schema(dfr_df: pd.DataFrame) -> pd.DataFrame:
    """Map normalized DFR records to DFR Summons Log column schema.

    Returns a DataFrame whose column names match the DFR Summons Log headers.
    Issuing Officer is set to the same value as DFR Operator (drone operator
    issues the ticket).
    """
    rows: list[dict] = []
    for _, row in dfr_df.iterrows():
        mapped: dict = {}
        for etl_col, dfr_col in _ETL_TO_DFR.items():
            val = row.get(etl_col, "")
            if pd.isna(val):
                val = ""
            mapped[dfr_col] = str(val).strip() if val != "" else ""

        # Issuing Officer mirrors DFR Operator for drone enforcement
        mapped["Issuing Officer"] = mapped.get("DFR Operator", "")

        # Leave manual-entry columns empty (filled by operator after review)
        mapped.setdefault("OCA", "")
        mapped.setdefault("Notes", "")

        rows.append(mapped)

    return pd.DataFrame(rows)


def export_to_dfr_workbook(dfr_df: pd.DataFrame, workbook_path: Path) -> bool:
    """Append new DFR summons rows to dfr_directed_patrol_enforcement.xlsx.

    - Deduplicates on 'Summons Number': rows already present are skipped.
    - Skips formula columns (DFR_FORMULA_COLUMN_NAMES) — Excel populates those.
    - On PermissionError (file open in Excel) saves a side-car `.etl_temp_dfr.xlsx`.

    Args:
        dfr_df:        DataFrame of DFR records from split_dfr_records().
        workbook_path: Absolute path to dfr_directed_patrol_enforcement.xlsx.

    Returns:
        True if records were written (or nothing to add), False on error.
    """
    if dfr_df.empty:
        logger.info("DFR export: no DFR records to write.")
        return True

    if not workbook_path.exists():
        logger.warning(
            "DFR workbook not found at %s — skipping export. "
            "Ensure the file exists at the configured OneDrive path.",
            workbook_path,
        )
        return False

    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        logger.error("openpyxl is not installed — cannot export to DFR workbook.")
        return False

    mapped_df = _map_to_dfr_schema(dfr_df)

    # ------------------------------------------------------------------ load
    try:
        wb = openpyxl.load_workbook(workbook_path)
    except Exception as exc:
        logger.error("Failed to open DFR workbook %s: %s", workbook_path, exc)
        return False

    # Locate the target sheet
    sheet = None
    if "DFR Summons Log" in wb.sheetnames:
        sheet = wb["DFR Summons Log"]
    else:
        # Fallback: first visible sheet that is not Instructions / ViolationData
        skip = {"Instructions", "ViolationData", "M Code Reference"}
        for name in wb.sheetnames:
            if name not in skip:
                sheet = wb[name]
                break
    if sheet is None:
        logger.error("Could not locate DFR Summons Log sheet in %s", workbook_path.name)
        wb.close()
        return False

    # Read header row (row 1) and build column-index map (0-based)
    header_row: list[str | None] = [cell.value for cell in sheet[1]]
    header_map: dict[str, int] = {
        str(name).strip(): idx
        for idx, name in enumerate(header_row)
        if name is not None
    }

    # --------------------------------------------------- existing summons IDs
    summons_col_idx = header_map.get("Summons Number")
    existing_numbers: set[str] = set()
    if summons_col_idx is not None:
        for row_cells in sheet.iter_rows(min_row=2, values_only=True):
            val = row_cells[summons_col_idx] if summons_col_idx < len(row_cells) else None
            if val not in (None, ""):
                existing_numbers.add(str(val).strip())

    # -------------------------------------------------------- append new rows
    added = 0
    skipped = 0
    for _, mapped_row in mapped_df.iterrows():
        ticket = str(mapped_row.get("Summons Number", "")).strip()
        if ticket and ticket in existing_numbers:
            skipped += 1
            continue

        # Build a new row aligned to workbook column positions
        new_row: list = [None] * len(header_row)
        for col_name, value in mapped_row.items():
            if col_name in DFR_FORMULA_COLUMN_NAMES:
                continue  # let Excel's formula populate this cell
            col_idx = header_map.get(col_name)
            if col_idx is not None:
                new_row[col_idx] = value if value != "" else None

        sheet.append(new_row)
        if ticket:
            existing_numbers.add(ticket)
        added += 1

    if added == 0:
        logger.info(
            "DFR workbook: all %d record(s) already present — nothing to add. (%d skipped)",
            skipped,
            skipped,
        )
        return True

    # --------------------------------------------------------------- save
    try:
        wb.save(workbook_path)
        logger.info(
            "DFR workbook: appended %d new record(s) to %s (%d duplicate(s) skipped).",
            added,
            workbook_path.name,
            skipped,
        )
        return True
    except PermissionError:
        # File is open in Excel — write to a side-car temp file for manual merge
        temp_path = workbook_path.parent / ".etl_temp_dfr.xlsx"
        try:
            wb.save(temp_path)
            logger.warning(
                "DFR workbook is open in Excel (PermissionError). "
                "Saved %d record(s) to side-car file: %s — merge manually.",
                added,
                temp_path.name,
            )
        except Exception as exc2:
            logger.error("Could not save DFR temp file either: %s", exc2)
        return False
    except Exception as exc:
        logger.error("Failed to save DFR workbook: %s", exc)
        return False
